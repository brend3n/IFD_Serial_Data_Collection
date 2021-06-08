# import requests
# from bs4 import BeautifulSoup
import time
import re
import csv
from openpyxl import Workbook, load_workbook

def save_to_file(course_dictionary, course_to_pre_req):
    # Save each dictionary in a different sheet

    wb = Workbook()
    
    ws_course_dictionary = wb.create_sheet("Course_Dictionary")
    ws_course_dictionary.title = "Course_Dictionary"

    ws_course_pre_reqs = wb.create_sheet("Course_to_Pre_Reqs")
    ws_course_pre_reqs.title = "Course_to_Pre_Reqs"

    row = 1
    col = 1
    
    # Fill data in Course_Dictionary
    for course in course_dictionary:
        # Write course in first column
        ws_course_dictionary.cell(column=col, row=row).value = course

        # Write name in next column
        col += 1
        ws_course_dictionary.cell(column=col, row=row).value = course_dictionary[course]

        # Move to next row
        col = 1
        row +=1

    row = 1
    col = 1
    # Fill data in Course_to_Pre_Reqs
    for course in course_to_pre_req:

        # Write course in first column
        ws_course_pre_reqs.cell(column=col, row=row).value = course

        # Set next column
        col += 1

        for pre_req in course_to_pre_req[course]:
            # Write each pre_req in the next column
            ws_course_pre_reqs.cell(column=col, row=row).value = pre_req
            col += 1

        # Reset column
        col = 1
        # Move to the next row
        row += 1

    wb.save("Course_Data.xlsx")

# Read the .xlsx file of scraped data
def read_from_file(file="Course_Data.xlsx"):

    # Declare dictionaries
    course_dictionary = {}
    course_to_pre_req = {}

    # Load in the workbook
    wb = load_workbook(file)

    # Get the worksheets in the spreadsheet
    ws_course_dictionary = wb["Course_Dictionary"]
    ws_course_pre_reqs = wb["Course_to_Pre_Reqs"]

    # Max rows in course_dictionary Excel sheet
    max_rows_course_dictionary = ws_course_dictionary.max_row

    # Extract the course_number and course name from the sheet and 
    # insert the mapping into the dictionary
    for row in range(1,max_rows_course_dictionary+1):

        # Get the course_number and course_name
        course_number = ws_course_dictionary.cell(row=row, column=1).value
        course_name = ws_course_dictionary.cell(row=row, column=2).value
        # Takes into account courses that go by multiple course numbers
        if "/" in course_number:
            prefix = course_number[:4]
            # print(f"prefix: <{prefix}>")
            course_number_variations = course_number[4:]
            course_number_variations = course_number_variations.split("/")
            # print(f"Course_number_variations: {course_number_variations}")
            for variation in course_number_variations:
                course_dictionary[f'{prefix}{variation}'] = course_name

        # Store mapping in dictionary
        course_dictionary[course_number] = course_name

    # Max rows in course_pre_reqs Excel sheet
    max_row_course_pre_reqs = ws_course_pre_reqs.max_row

    # Extract the course_number and the list of pre_req classes for that
    # course. Insert the mapping into the dictionary
    for row in range(1, max_row_course_pre_reqs+1):

        # Get the course number
        course_number = ws_course_pre_reqs.cell(row=row, column=1).value

        temp_list = []
        col = 2
        # print(f'res: {ws_course_pre_reqs.cell(row=4, column=2).value}')
        
        # Get first column element
        val = ws_course_pre_reqs.cell(row=row, column=col).value
        # Get the remainder of the elements if they exist
        while val != None:
            temp_list.append(val)
            col += 1
            val = ws_course_pre_reqs.cell(row=row, column=col).value
        
        course_to_pre_req[course_number] = temp_list


        # print(f"{course_number}:{temp_list}")
        
    return course_dictionary, course_to_pre_req

# Re_scrape all data and save to file to account for changes in courses and
# pre-reqs.
def update_file(file="Course_Data.xlsx"):
    course_dictionary = {}
    course_to_pre_req = {}

    # Last page number used to iterate through all pages on site
    last_page = get_last_page_number(default_url)

    # Loop through all pages
    for i in range(1,last_page + 1):

        # Get the page i
        url = url_generator(i)

        # Do something with the page
        res = get_course_list(url)

        # Merge dictionaries
        course_dictionary = {**course_dictionary,**res[0]}
        course_to_pre_req = {**course_to_pre_req,**res[1]}
 
    '''
    # Uncomment for printing
    print("Course_dictionary")
    for ele in course_dictionary:
        print(f"{ele} : {course_dictionary[ele]}\n")
    
    print("\n")
    print("Course_to_pre_req")
    for ele in course_to_pre_req:
        print(f"{ele} : {course_to_pre_req[ele]}")
    '''

    save_to_file(course_dictionary=course_dictionary, course_to_pre_req=course_to_pre_req)

    return course_dictionary, course_to_pre_req