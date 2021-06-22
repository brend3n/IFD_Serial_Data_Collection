'''
    Author: Brenden Morton
    Date Started: 6/05/2021
    Description: Test code for validating SD cards data corruption protection.
'''

#################################################
# Interfacing with powersupply (PSU)
import serial

# Interfacing with the IFD over CAN bus
#import can

# Timeing intervals
import time

# Writing data to excel spreadsheet
from openpyxl import Workbook

# File system/directories
import os  

# Obtaining current date when a test begins
from datetime import date
##################################################


# Iterations for testing
num_iterations = 10000

# (seconds) Minimum delay in between cycling power 
min_delay = 30

# (seconds) Maximum delay in between cycling power 
max_delay = 150

# Using the minimum delay for testing intially
delay = min_delay

# Baudrate for the device (default)
baud = 109200

# Port for the device (default)
# port = 'COM1'

# Date
current_date = date.today()

# Instantiate a workbook object for storing data
wb = Workbook()

# Creating a sheet in the workbook
test_data = wb.create_sheet(f"Test Data ({current_date})")

# Initializes the workbook for data storage
def init_wb():

    # Start in the first column
    col = 1
    
    # Write test number
    test_data.cell(column=col, row=1).value = 'test_no'
    
    # Next column
    col += 1

    # Write test result    
    test_data.cell(column=col, row=1).value = 'result'

    # Next column
    col +=1 

    # Write test data
    test_data.cell(column=col, row=1).value = 'test_data'

    # wb.remove("sheet")
    wb.save("SD_TEST_DATA.xlsx")

    return

# Records the test data in an Excel spreadsheet
def record_test(test_no, test_data, result):

    # Start in the first column
    col = 1
    
    # Write test number
    test_data.cell(column=col, row=(test_no + 1)).value = test_no
    
    # Next column
    col += 1

    # Write test result    
    test_data.cell(column=col, row=(test_no + 1)).value = result

    # Next column
    col +=1 

    # Write test data
    test_data.cell(column=col, row=(test_no + 1)).value = test_data

    # Saving the workbook
    wb.save("SD_TEST_DATA.xlsx")

    return


# TODO: Determine whether or not a fail occurs
# Returns: True if Pass and False if Fail
def parse_test(file_ref):

    # Assume pass
    res = True

    # TODO: Find fail

    return res, test_data


# Parse and evaluate the result of a test
def assert_test(test_no, file):

    # TODO: Parse test
    # !

    result, test_data = parse_test(file)

    if result:
        result = 'FAIL'
    elif not result:
        result = 'PASS'

    record_test(test_no, test_data, result)


# Loop to find the port
def loop():
    for i in range(1,100):
        port = f"COM{i}"
        print(f"COM {i}")
        # run_2(port) 
        time.sleep(3)

# Works for testing
def power_cycle(ps, time_delay, file):
    init_wb()
    iteration = 0
    while True:

        try:
            ps.open()
            print("Port opened!")
        except:
            print("Could not open port -> Exception thrown")
            ps.close()
        
        for i in range(time_delay):
            print(i)
            time.sleep(1)
        

        
        assert_test(iteration, file)
        iteration += 1

def run_test(file):

    # Initiate the connection to the port for Power Supply    
    ps = serial.Serial(
        port='COM3',
        baudrate=9600,
        parity=serial.PARITY_NONE,
        stopbits=serial.STOPBITS_ONE,
        bytesize=serial.EIGHTBITS,
        rtscts=True,
        timeout=0
    )
    # PSU is initially OFF
    ps.close()

    # Initiate the connection to the port of the IFD
    ifd = serial.Serial(
        port='COM4',
        baudrate=9600,
        parity=serial.PARITY_NONE,
        stopbits=serial.STOPBITS_ONE,
        bytesize=serial.EIGHTBITS,
        rtscts=True,
        timeout=0
    )

    while True:
        print(f"\n~Enter~\n[1] Turn on\n[2] Turn off\n[3] Power Cycle\n[4] To end program")
        res = input()
        res = int(res)

        # Evaluating the input of the menu.


        if res == 1:
            try:
                ps.open()
                os.system('cls')
                print("Turning on")
                
            except Exception as e:
                print(f"Exception: {e}")
                pass
        elif res == 2:
            try:
                ps.close()
                os.system('cls')
                print("Turning off")
            except Exception as e:
                print(f"Exception: {e}")
                pass
        elif res == 3:
            os.system('cls')
            print("Power cycle")
            delay = int(input("Enter delay in seconds: "))
            power_cycle(ps, delay, file)
        else:
            print("Ending program")
            return
              
def main(): 
    # TODO: Need to figure output file name and extension
    file = ""
    init_wb()
    run_test(file)



if __name__ == "__main__":
    main()
